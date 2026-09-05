import os
import re
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from urllib.parse import parse_qs, urljoin, urlparse

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry


# ============================================================
# CONFIG
# ============================================================

RACECARD_BASE_URL = (
    "https://racing.hkjc.com/en-us/local/information/racecard"
)

RACE_DATE = os.getenv(
    "RACE_DATE",
    "2026-09-06"
)

RACE_WORKERS = max(
    1,
    int(
        os.getenv(
            "RACE_WORKERS",
            "8"
        )
    )
)

HORSE_WORKERS = max(
    1,
    int(
        os.getenv(
            "HORSE_WORKERS",
            "20"
        )
    )
)

REQUEST_TIMEOUT = float(
    os.getenv(
        "REQUEST_TIMEOUT",
        "60"
    )
)

HTTP_RETRIES = max(
    0,
    int(
        os.getenv(
            "HTTP_RETRIES",
            "4"
        )
    )
)

HTTP_BACKOFF_FACTOR = float(
    os.getenv(
        "HTTP_BACKOFF_FACTOR",
        "1.5"
    )
)


# ============================================================
# OUTPUT FILES
# ============================================================

RESULTS_DIR = os.path.join(
    "results",
    "races"
)

OUTPUT_XLSX = os.path.join(
    RESULTS_DIR,
    "upcoming_races.xlsx"
)

OUTPUT_CSV = os.path.join(
    RESULTS_DIR,
    "upcoming_races.csv"
)


# ============================================================
# OUTPUT COLUMNS
# ============================================================

OUTPUT_COLUMNS = [

    # --------------------------------------------------------
    # RACE HEADER
    # --------------------------------------------------------

    "race_date",
    "racecourse_code",
    "racecourse_name",
    "race_number",
    "race_id",
    "race_name",
    "race_time",

    "surface",
    "course",
    "distance_m",
    "going",

    "prize_money_hkd",
    "rating_band",
    "race_class",

    # --------------------------------------------------------
    # RACECARD HORSE TABLE
    # --------------------------------------------------------

    "horse_number",
    "horse_name",
    "brand_number",

    "handicap_weight",
    "jockey",
    "probable_overweight",
    "draw",
    "trainer",

    "international_rating",
    "horse_rating",

    "declared_horse_weight",
    "horse_weight_change",

    "best_time",
    "wfa",
    "days_since_last_run",

    "owner",
    "import_category",

    # --------------------------------------------------------
    # HORSE PROFILE
    #
    # These are specifically taken from the horse page.
    # --------------------------------------------------------

    "horse_age",
    "horse_sex",

    "season_stakes_hkd",
    "total_stakes_hkd",

    "sire",
    "dam",

    # --------------------------------------------------------
    # IDS + SOURCE URLS
    # --------------------------------------------------------

    "horse_id",
    "runner_id",

    "horse_profile_url",
    "race_url",
]


# ============================================================
# HTTP
# ============================================================

COMMON_HEADERS = {

    "User-Agent": (
        "Mozilla/5.0 "
        "(Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 "
        "(KHTML, like Gecko) "
        "Chrome/140.0.0.0 "
        "Safari/537.36"
    ),

    "Accept": (
        "text/html,"
        "application/xhtml+xml,"
        "application/xml;q=0.9,"
        "*/*;q=0.8"
    ),

    "Accept-Language":
        "en-US,en;q=0.9",
}


_thread_local = threading.local()


def create_http_session():

    session = requests.Session()

    session.headers.update(
        COMMON_HEADERS
    )

    retry = Retry(

        total=
            HTTP_RETRIES,

        connect=
            HTTP_RETRIES,

        read=
            HTTP_RETRIES,

        status=
            HTTP_RETRIES,

        backoff_factor=
            HTTP_BACKOFF_FACTOR,

        status_forcelist=(
            429,
            500,
            502,
            503,
            504,
        ),

        allowed_methods=
            frozenset(
                [
                    "GET"
                ]
            ),

        respect_retry_after_header=
            True,

        raise_on_status=
            False,
    )

    adapter = HTTPAdapter(

        max_retries=
            retry,

        pool_connections=
            2,

        pool_maxsize=
            2,
    )

    session.mount(
        "https://",
        adapter
    )

    session.mount(
        "http://",
        adapter
    )

    return session


def get_session():

    session = getattr(
        _thread_local,
        "session",
        None
    )

    if session is None:

        session = (
            create_http_session()
        )

        _thread_local.session = (
            session
        )

    return session


def http_get(
    url
):

    try:

        response = (
            get_session()
            .get(
                url,

                timeout=
                    REQUEST_TIMEOUT,

                allow_redirects=
                    True,
            )
        )

        print(
            f"GET "
            f"{url} "
            f"-> "
            f"{response.status_code}"
        )

        response.raise_for_status()

        return response

    except requests.RequestException as exc:

        print(
            f"REQUEST FAILED "
            f"{url}: "
            f"{exc}"
        )

        return None


# ============================================================
# GENERAL HELPERS
# ============================================================

def clean_text(
    value
):

    if value is None:
        return ""

    return re.sub(
        r"\s+",
        " ",
        str(
            value
        ).replace(
            "\xa0",
            " "
        )
    ).strip()


def parse_int(
    value
):

    text = clean_text(
        value
    ).replace(
        ",",
        ""
    )

    match = re.search(
        r"-?\d+",
        text
    )

    if not match:
        return None

    try:

        return int(
            match.group()
        )

    except ValueError:

        return None


def parse_money(
    value
):

    text = clean_text(
        value
    )

    match = re.search(
        r"(?:HK\s*)?"
        r"\$\s*"
        r"([\d,]+)",

        text,
        re.I
    )

    if not match:
        return None

    try:

        return int(
            match.group(
                1
            ).replace(
                ",",
                ""
            )
        )

    except ValueError:

        return None


def normalise_header(
    value
):

    text = clean_text(
        value
    ).lower()

    text = text.replace(
        "’",
        "'"
    )

    text = text.replace(
        "+/-",
        " vs "
    )

    text = text.replace(
        "+",
        " "
    )

    text = text.replace(
        "-",
        " "
    )

    text = text.replace(
        "&",
        " and "
    )

    text = re.sub(
        r"[().'\"/\\]",
        " ",
        text
    )

    text = re.sub(
        r"[^a-z0-9 ]+",
        " ",
        text
    )

    return re.sub(
        r"\s+",
        " ",
        text
    ).strip()


def parse_requested_date():

    try:

        return datetime.strptime(
            RACE_DATE,
            "%Y-%m-%d"
        ).date()

    except ValueError as exc:

        raise SystemExit(
            "RACE_DATE must use "
            "YYYY-MM-DD, "
            "for example 2026-09-06"
        ) from exc


def requested_date_slash():

    return (
        parse_requested_date()
        .strftime(
            "%Y/%m/%d"
        )
    )


def build_date_url():

    return (
        f"{RACECARD_BASE_URL}"
        f"?racedate="
        f"{requested_date_slash()}"
    )


def build_race_url(
    racecourse_code,
    race_number
):

    return (
        f"{RACECARD_BASE_URL}"
        f"?racedate="
        f"{requested_date_slash()}"
        f"&Racecourse="
        f"{racecourse_code}"
        f"&RaceNo="
        f"{race_number}"
    )


def query_value_case_insensitive(
    query,
    key
):

    for (
        existing_key,
        values
    ) in query.items():

        if (
            existing_key.lower()
            ==
            key.lower()
            and
            values
        ):

            return values[
                0
            ]

    return ""


def extract_horse_id(
    url
):

    if not url:
        return ""

    query = parse_qs(
        urlparse(
            url
        ).query
    )

    return clean_text(
        query_value_case_insensitive(
            query,
            "horseid"
        )
    )


# ============================================================
# DISCOVER ALL LOCAL HK RACES ON REQUESTED DATE
# ============================================================

def discover_races():

    date_url = (
        build_date_url()
    )

    response = http_get(
        date_url
    )

    if response is None:
        return []

    soup = BeautifulSoup(
        response.text,
        "html.parser"
    )

    requested = (
        requested_date_slash()
    )

    tasks = {}

    # --------------------------------------------------------
    # Look for the actual HKJC race navigation links.
    # --------------------------------------------------------

    for link in soup.find_all(
        "a",
        href=True
    ):

        href = urljoin(
            response.url,
            link.get(
                "href",
                ""
            )
        )

        parsed = urlparse(
            href
        )

        if (
            "racecard"
            not in
            parsed.path.lower()
        ):

            continue

        query = parse_qs(
            parsed.query
        )

        race_no_text = (
            query_value_case_insensitive(
                query,
                "RaceNo"
            )
        )

        racecourse = (
            query_value_case_insensitive(
                query,
                "Racecourse"
            )
            .upper()
        )

        race_date = (
            query_value_case_insensitive(
                query,
                "racedate"
            )
        )

        race_no = parse_int(
            race_no_text
        )

        if race_no is None:
            continue

        if racecourse not in {
            "ST",
            "HV",
        }:
            continue

        if (
            race_date
            and
            race_date
            !=
            requested
        ):
            continue

        tasks[
            (
                racecourse,
                race_no
            )
        ] = {

            "racecourse_code":
                racecourse,

            "race_number":
                race_no,

            "race_url":
                build_race_url(
                    racecourse,
                    race_no
                ),
        }

    # --------------------------------------------------------
    # FALLBACK
    #
    # If HKJC changes the race navigation links, determine
    # venue from page text and probe R1-R12.
    #
    # Any race number with no table is simply ignored.
    # --------------------------------------------------------

    if not tasks:

        page_text = clean_text(
            soup.get_text(
                " ",
                strip=True
            )
        )

        if (
            "Sha Tin"
            in
            page_text
        ):

            racecourse = "ST"

        elif (
            "Happy Valley"
            in
            page_text
        ):

            racecourse = "HV"

        else:

            print(
                "No local Hong Kong "
                "race meeting found "
                "for this date."
            )

            return []

        for race_no in range(
            1,
            13
        ):

            tasks[
                (
                    racecourse,
                    race_no
                )
            ] = {

                "racecourse_code":
                    racecourse,

                "race_number":
                    race_no,

                "race_url":
                    build_race_url(
                        racecourse,
                        race_no
                    ),
            }

    discovered = sorted(

        tasks.values(),

        key=lambda item:
            (
                item[
                    "racecourse_code"
                ],

                item[
                    "race_number"
                ]
            )
    )

    print(
        f"Discovered/queued "
        f"{len(discovered)} "
        f"local race pages."
    )

    return discovered


# ============================================================
# RACE HEADER
# ============================================================

def parse_race_header(
    soup,
    racecourse_code,
    fallback_race_number,
    race_url
):

    strings = [

        clean_text(
            item
        )

        for item in
        soup.stripped_strings

        if clean_text(
            item
        )
    ]

    race_number = (
        fallback_race_number
    )

    race_name = ""
    race_time = ""

    racecourse_name = (
        "Sha Tin"
        if racecourse_code == "ST"
        else
        "Happy Valley"
    )

    surface = ""
    course = ""
    distance_m = None
    going = ""

    prize_money_hkd = None
    rating_band = ""
    race_class = ""

    race_heading_index = None

    # --------------------------------------------------------
    # Example:
    #
    # Race 3 - THE HKSAR CHIEF EXECUTIVE'S CUP (HANDICAP)
    # --------------------------------------------------------

    for index, text in enumerate(
        strings
    ):

        match = re.match(
            r"^Race\s+"
            r"(\d+)"
            r"\s*-\s*"
            r"(.+)$",

            text,
            re.I
        )

        if match:

            race_number = int(
                match.group(
                    1
                )
            )

            race_name = clean_text(
                match.group(
                    2
                )
            )

            race_heading_index = (
                index
            )

            break

    # Only inspect strings close to the race title so unrelated
    # navigation content cannot be mistaken for race metadata.

    if race_heading_index is not None:

        nearby = strings[
            race_heading_index:
            race_heading_index + 15
        ]

    else:

        nearby = strings

    for text in nearby:

        # ----------------------------------------------------
        # Example:
        #
        # Sunday, September 06, 2026, Sha Tin, 13:30
        # ----------------------------------------------------

        if (
            (
                "Sha Tin"
                in
                text
                or
                "Happy Valley"
                in
                text
            )
            and
            re.search(
                r"\b\d{1,2}:\d{2}\b",
                text
            )
        ):

            if (
                "Sha Tin"
                in
                text
            ):

                racecourse_name = (
                    "Sha Tin"
                )

            else:

                racecourse_name = (
                    "Happy Valley"
                )

            time_match = re.search(
                r"\b"
                r"(\d{1,2}:\d{2})"
                r"\b",

                text
            )

            if time_match:

                race_time = (
                    time_match.group(
                        1
                    )
                )

        # ----------------------------------------------------
        # Example:
        #
        # Turf, "A" Course, 1200M, Good
        # ----------------------------------------------------

        distance_match = re.search(
            r"\b"
            r"(\d{3,4})"
            r"\s*M\b",

            text,
            re.I
        )

        if (
            distance_match
            and
            (
                "Turf"
                in
                text
                or
                "All Weather"
                in
                text
                or
                "AWT"
                in
                text
            )
        ):

            distance_m = int(
                distance_match.group(
                    1
                )
            )

            parts = [

                clean_text(
                    part
                )

                for part in
                text.split(
                    ","
                )
            ]

            if parts:

                surface = parts[
                    0
                ]

            course_match = re.search(
                r'["“]?'
                r'([^,"”]+)'
                r'["”]?'
                r'\s+Course',

                text,
                re.I
            )

            if course_match:

                course = clean_text(
                    course_match.group(
                        1
                    )
                ).strip(
                    '"“”'
                )

            if len(
                parts
            ) >= 2:

                going = parts[
                    -1
                ]

        # ----------------------------------------------------
        # Example:
        #
        # Prize Money: $4,200,000, -, Group Three
        #
        # or
        #
        # Prize Money: $1,170,000, Rating 40-0, Class 4
        # ----------------------------------------------------

        if (
            "Prize Money"
            in
            text
        ):

            prize_money_hkd = (
                parse_money(
                    text
                )
            )

            rating_match = re.search(
                r"Rating\s+"
                r"([A-Za-z0-9+\-\s]+?)"
                r"(?=,|$)",

                text,
                re.I
            )

            if rating_match:

                rating_band = clean_text(
                    rating_match.group(
                        1
                    )
                )

            class_patterns = [

                r"\bClass\s+\d+\b",

                r"\bGroup\s+"
                r"(?:One|Two|Three|1|2|3)"
                r"\b",

                r"\bGriffin\b",

                r"\bListed\b",
            ]

            for pattern in (
                class_patterns
            ):

                class_match = re.search(
                    pattern,
                    text,
                    re.I
                )

                if class_match:

                    race_class = clean_text(
                        class_match.group(
                            0
                        )
                    ).title()

                    break

            if not rating_band:

                for part in [

                    clean_text(
                        item
                    )

                    for item in
                    text.split(
                        ","
                    )
                ]:

                    if re.fullmatch(
                        r"\d+\s*-\s*\d+",
                        part
                    ):

                        rating_band = (
                            part
                        )

                        break

    race_date = (
        parse_requested_date()
        .strftime(
            "%Y-%m-%d"
        )
    )

    race_id = (
        f"{racecourse_code}_"
        f"{parse_requested_date().strftime('%Y%m%d')}_"
        f"R{race_number:02d}"
    )

    return {

        "race_date":
            race_date,

        "racecourse_code":
            racecourse_code,

        "racecourse_name":
            racecourse_name,

        "race_number":
            race_number,

        "race_id":
            race_id,

        "race_name":
            race_name,

        "race_time":
            race_time,

        "surface":
            surface,

        "course":
            course,

        "distance_m":
            distance_m,

        "going":
            going,

        "prize_money_hkd":
            prize_money_hkd,

        "rating_band":
            rating_band,

        "race_class":
            race_class,

        "race_url":
            race_url,
    }


# ============================================================
# RACECARD TABLE
# ============================================================

TABLE_HEADER_ALIASES = {

    "horse_number": {
        "horse no",
        "horse number",
    },

    "horse_name": {
        "horse",
        "name",
    },

    "brand_number": {
        "brand no",
        "brand number",
    },

    "handicap_weight": {
        "wt",
        "handicap weight",
    },

    "jockey": {
        "jockey",
    },

    "probable_overweight": {
        "over wt",
        "probable overweight",
    },

    "draw": {
        "draw",
        "dr",
    },

    "trainer": {
        "trainer",
    },

    "international_rating": {
        "intl rtg",
        "int l rtg",
        "international rating",
    },

    "horse_rating": {
        "rtg",
        "rating",
    },

    "declared_horse_weight": {
        "horse wt declaration",
        "horse wt declaration horse weight",
        "horse weight declaration",
    },

    "horse_weight_change": {
        "wt vs declaration",
        "wt vs vs declaration",
        "wt vs declaration horse weight",
        "wt vs declaration vs declaration",
    },

    "best_time": {
        "best time",
    },

    "wfa": {
        "wfa",
        "weight for age allowance",
    },

    "days_since_last_run": {
        "days since last run",
    },

    "owner": {
        "owner",
        "owners",
    },

    "import_category": {
        "import cat",
        "import category",
    },
}


# ============================================================
# THESE RACECARD COLUMNS ARE DELIBERATELY EXCLUDED:
#
# Last 6 Runs
# Colour
# Rtg.+/-
# Priority
# Gear
#
# Age
# Sex
# Season Stakes
# Sire
# Dam
#
# Age/Sex/Stakes/Sire/Dam are taken from the horse profile
# page instead.
# ============================================================


def identify_table_header(
    value
):

    normalised = (
        normalise_header(
            value
        )
    )

    if not normalised:
        return None

    for (
        canonical,
        aliases
    ) in (
        TABLE_HEADER_ALIASES
        .items()
    ):

        if (
            normalised
            in
            aliases
        ):

            return canonical

    return None


def build_column_map(
    table
):

    best_map = {}

    for row in table.find_all(
        "tr"
    ):

        cells = row.find_all(
            [
                "th",
                "td",
            ],
            recursive=False
        )

        if not cells:

            cells = row.find_all(
                [
                    "th",
                    "td",
                ]
            )

        current = {}

        for index, cell in enumerate(
            cells
        ):

            canonical = (
                identify_table_header(
                    cell.get_text(
                        " ",
                        strip=True
                    )
                )
            )

            if (
                canonical
                and
                canonical
                not in
                current
            ):

                current[
                    canonical
                ] = index

        required = {
            "horse_number",
            "horse_name",
            "jockey",
            "trainer",
        }

        if (
            required.issubset(
                current.keys()
            )
            and
            len(
                current
            )
            >
            len(
                best_map
            )
        ):

            best_map = (
                current
            )

    return best_map


def find_main_runner_table(
    soup
):

    best_table = None
    best_map = {}

    for table in soup.find_all(
        "table"
    ):

        column_map = (
            build_column_map(
                table
            )
        )

        if (
            len(
                column_map
            )
            >
            len(
                best_map
            )
        ):

            best_table = (
                table
            )

            best_map = (
                column_map
            )

    return (
        best_table,
        best_map
    )


def get_cell(
    cells,
    column_map,
    field
):

    index = (
        column_map.get(
            field
        )
    )

    if (
        index is None
        or
        index < 0
        or
        index >= len(
            cells
        )
    ):

        return None

    return cells[
        index
    ]


def get_cell_text(
    cells,
    column_map,
    field
):

    cell = get_cell(
        cells,
        column_map,
        field
    )

    if cell is None:
        return ""

    return clean_text(
        cell.get_text(
            " ",
            strip=True
        )
    )


def parse_runners(
    soup,
    race_header
):

    (
        table,
        column_map
    ) = find_main_runner_table(
        soup
    )

    if (
        table is None
        or
        not column_map
    ):

        return []

    runners = []

    for row in table.find_all(
        "tr"
    ):

        cells = row.find_all(
            "td",
            recursive=False
        )

        if not cells:

            cells = row.find_all(
                "td"
            )

        if not cells:
            continue

        horse_number = parse_int(
            get_cell_text(
                cells,
                column_map,
                "horse_number"
            )
        )

        horse_cell = get_cell(
            cells,
            column_map,
            "horse_name"
        )

        horse_link = None

        if horse_cell is not None:

            horse_link = (
                horse_cell.find(
                    "a",

                    href=
                        re.compile(
                            r"horse\?horseid=",
                            re.I
                        )
                )
            )

        if horse_link is None:

            horse_link = row.find(
                "a",

                href=
                    re.compile(
                        r"horse\?horseid=",
                        re.I
                    )
            )

        if (
            horse_number is None
            or
            horse_link is None
        ):

            continue

        horse_profile_url = (
            urljoin(
                "https://racing.hkjc.com",

                horse_link.get(
                    "href",
                    ""
                )
            )
        )

        horse_id = extract_horse_id(
            horse_profile_url
        )

        horse_name = clean_text(
            horse_link.get_text(
                " ",
                strip=True
            )
        )

        runner = dict(
            race_header
        )

        runner.update({

            "horse_number":
                horse_number,

            "horse_name":
                horse_name,

            "brand_number":
                get_cell_text(
                    cells,
                    column_map,
                    "brand_number"
                ),

            "handicap_weight":
                parse_int(
                    get_cell_text(
                        cells,
                        column_map,
                        "handicap_weight"
                    )
                ),

            "jockey":
                get_cell_text(
                    cells,
                    column_map,
                    "jockey"
                ),

            "probable_overweight":
                parse_int(
                    get_cell_text(
                        cells,
                        column_map,
                        "probable_overweight"
                    )
                ),

            "draw":
                parse_int(
                    get_cell_text(
                        cells,
                        column_map,
                        "draw"
                    )
                ),

            "trainer":
                get_cell_text(
                    cells,
                    column_map,
                    "trainer"
                ),

            "international_rating":
                parse_int(
                    get_cell_text(
                        cells,
                        column_map,
                        "international_rating"
                    )
                ),

            "horse_rating":
                parse_int(
                    get_cell_text(
                        cells,
                        column_map,
                        "horse_rating"
                    )
                ),

            "declared_horse_weight":
                parse_int(
                    get_cell_text(
                        cells,
                        column_map,
                        "declared_horse_weight"
                    )
                ),

            "horse_weight_change":
                parse_int(
                    get_cell_text(
                        cells,
                        column_map,
                        "horse_weight_change"
                    )
                ),

            "best_time":
                get_cell_text(
                    cells,
                    column_map,
                    "best_time"
                ),

            "wfa":
                get_cell_text(
                    cells,
                    column_map,
                    "wfa"
                ),

            "days_since_last_run":
                parse_int(
                    get_cell_text(
                        cells,
                        column_map,
                        "days_since_last_run"
                    )
                ),

            "owner":
                get_cell_text(
                    cells,
                    column_map,
                    "owner"
                ),

            "import_category":
                get_cell_text(
                    cells,
                    column_map,
                    "import_category"
                ),

            # Filled from horse profile later.
            "horse_age":
                None,

            "horse_sex":
                "",

            "season_stakes_hkd":
                None,

            "total_stakes_hkd":
                None,

            "sire":
                "",

            "dam":
                "",

            "horse_id":
                horse_id,

            "runner_id":
                (
                    f"{race_header['race_id']}_"
                    f"{horse_number:02d}"
                ),

            "horse_profile_url":
                horse_profile_url,
        })

        runners.append(
            runner
        )

    return runners


# ============================================================
# HORSE PROFILE
#
# ONLY:
#
# age
# sex
# season stakes
# total stakes
# sire
# dam
# ============================================================

def extract_horse_profile_fields(
    html,
    horse_id,
    profile_url
):

    soup = BeautifulSoup(
        html,
        "html.parser"
    )

    text = clean_text(
        soup.get_text(
            " ",
            strip=True
        )
    )

    age = None
    sex = ""

    season_stakes_hkd = None
    total_stakes_hkd = None

    sire = ""
    dam = ""

    # --------------------------------------------------------
    # Country of Origin / Age : NZ / 6
    # --------------------------------------------------------

    origin_age = re.search(
        r"Country of Origin"
        r"\s*/\s*"
        r"Age"
        r"\s*:\s*"
        r"(.*?)"
        r"\s+"
        r"Colour"
        r"\s*/\s*"
        r"Sex"
        r"\s*:",

        text,
        re.I
    )

    if origin_age:

        age_match = re.search(
            r"/\s*(\d+)\s*$",

            clean_text(
                origin_age.group(
                    1
                )
            )
        )

        if age_match:

            age = int(
                age_match.group(
                    1
                )
            )

    # --------------------------------------------------------
    # Colour / Sex : Bay / Gelding
    #
    # We ignore colour and only keep sex.
    # --------------------------------------------------------

    colour_sex = re.search(
        r"Colour"
        r"\s*/\s*"
        r"Sex"
        r"\s*:\s*"
        r"(.*?)"
        r"\s+"
        r"Import Type"
        r"\s*:",

        text,
        re.I
    )

    if colour_sex:

        value = clean_text(
            colour_sex.group(
                1
            )
        )

        if "/" in value:

            sex = clean_text(
                value.rsplit(
                    "/",
                    1
                )[
                    1
                ]
            )

        else:

            sex = value

    # --------------------------------------------------------
    # Season Stakes* : $0
    # --------------------------------------------------------

    season_match = re.search(
        r"Season Stakes"
        r"\*?"
        r"\s*:\s*"
        r"((?:HK\s*)?"
        r"\$\s*"
        r"[\d,]+)",

        text,
        re.I
    )

    if season_match:

        season_stakes_hkd = (
            parse_money(
                season_match.group(
                    1
                )
            )
        )

    # --------------------------------------------------------
    # Total Stakes* : $158,571,900
    # --------------------------------------------------------

    total_match = re.search(
        r"Total Stakes"
        r"\*?"
        r"\s*:\s*"
        r"((?:HK\s*)?"
        r"\$\s*"
        r"[\d,]+)",

        text,
        re.I
    )

    if total_match:

        total_stakes_hkd = (
            parse_money(
                total_match.group(
                    1
                )
            )
        )

    # --------------------------------------------------------
    # Sire : Shamexpress
    # Dam : Missy Moo
    # Dam's Sire : Per Incanto
    # --------------------------------------------------------

    sire_match = re.search(
        r"\bSire"
        r"\s*:\s*"
        r"(.*?)"
        r"\s+Dam"
        r"\s*:",

        text,
        re.I
    )

    if sire_match:

        sire = clean_text(
            sire_match.group(
                1
            )
        )

    dam_match = re.search(
        r"\bDam"
        r"\s*:\s*"
        r"(.*?)"
        r"\s+Dam['’]s Sire"
        r"\s*:",

        text,
        re.I
    )

    if dam_match:

        dam = clean_text(
            dam_match.group(
                1
            )
        )

    return {

        "horse_id":
            horse_id,

        "horse_profile_url":
            profile_url,

        "horse_age":
            age,

        "horse_sex":
            sex,

        "season_stakes_hkd":
            season_stakes_hkd,

        "total_stakes_hkd":
            total_stakes_hkd,

        "sire":
            sire,

        "dam":
            dam,
    }


def fetch_horse_profile(
    horse_id,
    profile_url
):

    response = http_get(
        profile_url
    )

    if response is None:

        return (
            horse_id,
            None
        )

    try:

        profile = (
            extract_horse_profile_fields(
                response.text,
                horse_id,
                response.url
            )
        )

        return (
            horse_id,
            profile
        )

    except Exception as exc:

        print(
            f"HORSE PROFILE PARSE FAILED "
            f"{horse_id}: "
            f"{exc}"
        )

        return (
            horse_id,
            None
        )


# ============================================================
# FETCH ONE RACE
# ============================================================

def fetch_race(
    task
):

    response = http_get(
        task[
            "race_url"
        ]
    )

    if response is None:

        return (
            task,
            [],
            "request_failed"
        )

    soup = BeautifulSoup(
        response.text,
        "html.parser"
    )

    race_header = (
        parse_race_header(
            soup,

            task[
                "racecourse_code"
            ],

            task[
                "race_number"
            ],

            response.url
        )
    )

    runners = parse_runners(
        soup,
        race_header
    )

    # No table = not an error.
    # Could simply be a non-existent race number or the
    # racecard has not been published yet.

    if not runners:

        return (
            task,
            [],
            "no_runner_table"
        )

    return (
        task,
        runners,
        None
    )


# ============================================================
# FETCH ALL RACES CONCURRENTLY
# ============================================================

def fetch_all_races(
    tasks
):

    if not tasks:
        return []

    worker_count = min(
        RACE_WORKERS,
        len(
            tasks
        )
    )

    all_runners = []

    print()
    print(
        "=" * 70
    )

    print(
        "RACE PHASE"
    )

    print(
        "=" * 70
    )

    print(
        f"Race pages queued: "
        f"{len(tasks)}"
    )

    print(
        f"Concurrent race workers: "
        f"{worker_count}"
    )

    with ThreadPoolExecutor(
        max_workers=
            worker_count
    ) as executor:

        futures = {

            executor.submit(
                fetch_race,
                task
            ):
                task

            for task in
            tasks
        }

        for future in as_completed(
            futures
        ):

            task = futures[
                future
            ]

            try:

                (
                    returned_task,
                    runners,
                    status
                ) = future.result()

            except Exception as exc:

                print(
                    f"RACE WORKER FAILED "
                    f"{task['racecourse_code']} "
                    f"R{task['race_number']}: "
                    f"{exc}"
                )

                continue

            if (
                status
                ==
                "no_runner_table"
            ):

                print(
                    f"SKIP NO RACECARD TABLE: "
                    f"{returned_task['racecourse_code']} "
                    f"R{returned_task['race_number']}"
                )

                continue

            if status is not None:

                print(
                    f"RACE FAILED: "
                    f"{returned_task['racecourse_code']} "
                    f"R{returned_task['race_number']} "
                    f"{status}"
                )

                continue

            print(
                f"RACE OK: "
                f"{returned_task['racecourse_code']} "
                f"R{returned_task['race_number']} "
                f"-> "
                f"{len(runners)} runners"
            )

            all_runners.extend(
                runners
            )

    return all_runners


# ============================================================
# FETCH UNIQUE HORSE PROFILES CONCURRENTLY
# ============================================================

def enrich_horses(
    runners
):

    unique_horses = {}

    for runner in runners:

        horse_id = clean_text(
            runner.get(
                "horse_id"
            )
        )

        profile_url = clean_text(
            runner.get(
                "horse_profile_url"
            )
        )

        if (
            horse_id
            and
            profile_url
        ):

            unique_horses[
                horse_id
            ] = profile_url

    if not unique_horses:
        return runners

    worker_count = min(
        HORSE_WORKERS,
        len(
            unique_horses
        )
    )

    print()
    print(
        "=" * 70
    )

    print(
        "HORSE PROFILE PHASE"
    )

    print(
        "=" * 70
    )

    print(
        f"Unique horse pages: "
        f"{len(unique_horses)}"
    )

    print(
        f"Concurrent horse workers: "
        f"{worker_count}"
    )

    profiles = {}

    with ThreadPoolExecutor(
        max_workers=
            worker_count
    ) as executor:

        futures = {

            executor.submit(
                fetch_horse_profile,
                horse_id,
                profile_url
            ):
                horse_id

            for (
                horse_id,
                profile_url
            ) in (
                unique_horses.items()
            )
        }

        for future in as_completed(
            futures
        ):

            horse_id = futures[
                future
            ]

            try:

                (
                    returned_id,
                    profile
                ) = future.result()

            except Exception as exc:

                print(
                    f"HORSE WORKER FAILED "
                    f"{horse_id}: "
                    f"{exc}"
                )

                continue

            if (
                returned_id
                ==
                horse_id
                and
                profile
            ):

                profiles[
                    horse_id
                ] = profile

    # --------------------------------------------------------
    # MERGE PROFILE FIELDS BACK INTO EACH RACECARD RUNNER
    # --------------------------------------------------------

    for runner in runners:

        horse_id = clean_text(
            runner.get(
                "horse_id"
            )
        )

        profile = profiles.get(
            horse_id
        )

        if not profile:
            continue

        runner[
            "horse_age"
        ] = profile.get(
            "horse_age"
        )

        runner[
            "horse_sex"
        ] = profile.get(
            "horse_sex",
            ""
        )

        runner[
            "season_stakes_hkd"
        ] = profile.get(
            "season_stakes_hkd"
        )

        runner[
            "total_stakes_hkd"
        ] = profile.get(
            "total_stakes_hkd"
        )

        runner[
            "sire"
        ] = profile.get(
            "sire",
            ""
        )

        runner[
            "dam"
        ] = profile.get(
            "dam",
            ""
        )

        runner[
            "horse_profile_url"
        ] = profile.get(
            "horse_profile_url",

            runner.get(
                "horse_profile_url",
                ""
            )
        )

    return runners


# ============================================================
# SORT RESULTS
# ============================================================

def sort_runners(
    runners
):

    return sorted(

        runners,

        key=lambda row:
            (

                clean_text(
                    row.get(
                        "racecourse_code"
                    )
                ),

                (
                    row.get(
                        "race_number"
                    )

                    if (
                        row.get(
                            "race_number"
                        )
                        is not None
                    )

                    else 999
                ),

                (
                    row.get(
                        "horse_number"
                    )

                    if (
                        row.get(
                            "horse_number"
                        )
                        is not None
                    )

                    else 999
                ),
            )
    )


# ============================================================
# CSV
# ============================================================

def write_csv(
    runners
):

    df = pd.DataFrame(
        runners
    )

    for column in (
        OUTPUT_COLUMNS
    ):

        if column not in df.columns:

            df[
                column
            ] = None

    df = df[
        OUTPUT_COLUMNS
    ]

    df.to_csv(
        OUTPUT_CSV,
        index=False
    )


# ============================================================
# EXCEL
# ============================================================

def write_xlsx(
    runners
):

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = (
        "upcoming_races"
    )

    sheet.freeze_panes = (
        "A2"
    )

    sheet.sheet_view.showGridLines = (
        False
    )

    header_fill = PatternFill(
        "solid",
        fgColor="1F4E78"
    )

    header_font = Font(
        color="FFFFFF",
        bold=True
    )

    imported_font = Font(
        color="008000"
    )

    # --------------------------------------------------------
    # HEADER ROW
    # --------------------------------------------------------

    for (
        col_index,
        column
    ) in enumerate(
        OUTPUT_COLUMNS,
        start=1
    ):

        cell = sheet.cell(
            row=1,
            column=col_index,
            value=column
        )

        cell.fill = (
            header_fill
        )

        cell.font = (
            header_font
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    # --------------------------------------------------------
    # DATA
    # --------------------------------------------------------

    for (
        row_index,
        runner
    ) in enumerate(
        runners,
        start=2
    ):

        for (
            col_index,
            column
        ) in enumerate(
            OUTPUT_COLUMNS,
            start=1
        ):

            cell = sheet.cell(

                row=
                    row_index,

                column=
                    col_index,

                value=
                    runner.get(
                        column
                    )
            )

            cell.font = (
                imported_font
            )

            cell.alignment = (
                Alignment(
                    vertical="top"
                )
            )

            if column in {

                "prize_money_hkd",
                "season_stakes_hkd",
                "total_stakes_hkd",

            }:

                cell.number_format = (
                    '$#,##0'
                )

    # --------------------------------------------------------
    # FILTERS
    # --------------------------------------------------------

    last_row = max(
        1,
        sheet.max_row
    )

    last_col = len(
        OUTPUT_COLUMNS
    )

    sheet.auto_filter.ref = (
        f"A1:"
        f"{get_column_letter(last_col)}"
        f"{last_row}"
    )

    # --------------------------------------------------------
    # COLUMN WIDTHS
    # --------------------------------------------------------

    width_overrides = {

        "race_name":
            38,

        "horse_name":
            24,

        "jockey":
            18,

        "trainer":
            18,

        "owner":
            28,

        "sire":
            22,

        "dam":
            22,

        "horse_profile_url":
            45,

        "race_url":
            45,
    }

    for (
        col_index,
        column
    ) in enumerate(
        OUTPUT_COLUMNS,
        start=1
    ):

        width = (
            width_overrides.get(

                column,

                min(
                    max(
                        len(
                            column
                        )
                        +
                        2,

                        11
                    ),

                    20
                )
            )
        )

        sheet.column_dimensions[
            get_column_letter(
                col_index
            )
        ].width = width

    sheet.row_dimensions[
        1
    ].height = 28

    workbook.save(
        OUTPUT_XLSX
    )


# ============================================================
# WRITE OUTPUTS
# ============================================================

def write_outputs(
    runners
):

    os.makedirs(
        RESULTS_DIR,
        exist_ok=True
    )

    runners = sort_runners(
        runners
    )

    write_csv(
        runners
    )

    write_xlsx(
        runners
    )

    print()

    print(
        f"Saved "
        f"{len(runners)} "
        f"runner rows"
    )

    print(
        f"CSV:  "
        f"{OUTPUT_CSV}"
    )

    print(
        f"XLSX: "
        f"{OUTPUT_XLSX}"
    )


# ============================================================
# MAIN
# ============================================================

def main():

    race_date = (
        parse_requested_date()
    )

    print(
        "=" * 70
    )

    print(
        "HKJC UPCOMING "
        "RACECARD COLLECTOR"
    )

    print(
        "=" * 70
    )

    print(
        f"Race date: "
        f"{race_date}"
    )

    print(
        f"Race workers: "
        f"{RACE_WORKERS}"
    )

    print(
        f"Horse workers: "
        f"{HORSE_WORKERS}"
    )

    # ========================================================
    # STEP 1
    # DISCOVER ALL HK LOCAL RACES ON THAT DATE
    # ========================================================

    tasks = discover_races()

    if not tasks:

        print(
            "No local HKJC "
            "races/racecards found "
            "for the requested date."
        )

        write_outputs(
            []
        )

        return

    # ========================================================
    # STEP 2
    # FETCH EVERY RACE CONCURRENTLY
    # ========================================================

    runners = fetch_all_races(
        tasks
    )

    if not runners:

        print(
            "No published runner "
            "tables found for the "
            "requested date."
        )

        write_outputs(
            []
        )

        return

    # ========================================================
    # STEP 3
    # FETCH EACH UNIQUE HORSE PROFILE ONCE
    # ========================================================

    runners = enrich_horses(
        runners
    )

    # ========================================================
    # STEP 4
    # WRITE UPCOMING CSV + EXCEL
    # ========================================================

    write_outputs(
        runners
    )


if __name__ == "__main__":
    main()
